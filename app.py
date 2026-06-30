import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from io import BytesIO
import gspread
from datetime import datetime

# --- LOAD DATABASE PRODUK ---
@st.cache_data
def load_product_db():
    # Pastikan file prodname.xlsx ada di folder yang sama
    df = pd.read_excel("prodname.xlsx", dtype={'Barcode': str})
    
    # Menghapus spasi tambahan jika ada
    df['Barcode'] = df['Barcode'].str.strip()
    return df

product_db = load_product_db()

# --- SETUP ---
st.set_page_config(page_title="Pengajuan Voucher", layout="wide")

# --- DATA REGION & TOKO ---
data_toko = {
    "Region 1": ["Pasar Rebo", "Kelapa Gading", "Ciputat", "Alam Sutera", "Medan", "Palembang", "Pekannbaru", "Jatake", "Serang", "Batam", "Lampung", "Serpong"],
    "Region 2": ["Meruya", "Bandung", "Cibitung", "Bekasi", "Cikarang", "Cirebon", "Bogor", "Tasikmalaya", "Pakansari", "Kerawang", "Cimahi", "Tegal"],
    "Region 3": ["Sidoarjo", "Denpasar", "Semarang", "Makasar", "Yogyakarta", "Banjarmasin", "Solo", "Balikpapan", "Mastrip", "Samarinda", "Manado", "Mataram"]
}

if 'daftar_pengajuan' not in st.session_state: 
    st.session_state.daftar_pengajuan = []

# --- FUNGSI GOOGLE SHEETS ---
def simpan_ke_googlesheets(data_list, store_val): # Tambahkan parameter store_val
    try:
        creds_dict = dict(st.secrets["gcp"])
        client = gspread.service_account_from_dict(creds_dict)
        sheet = client.open_by_url("https://docs.google.com/spreadsheets/d/1qcixEUIKDJHebYxqqEHnZNDqhX_kEdKcbSfJx4guvwU/edit").sheet1
        today = datetime.now().strftime("%Y-%m-%d")
        
        data_exist = sheet.get_all_values()
        next_row = len(data_exist) + 1 
        
        for item in data_list:
            updates = [
                {'range': f'A{next_row}', 'values': [[next_row - 5]]},
                {'range': f'B{next_row}', 'values': [[store_val]]}, 
                {'range': f'C{next_row}', 'values': [[today]]},
                {'range': f'D{next_row}', 'values': [[item.get('No Cust')]]},
                {'range': f'E{next_row}', 'values': [[item.get('Nama')]]},
                {'range': f'F{next_row}', 'values': [[item.get('Barcode')]]},
                {'range': f'G{next_row}', 'values': [[item.get('Prodname')]]},
                {'range': f'H{next_row}', 'values': [[item.get('QTY')]]},
                {'range': f'I{next_row}', 'values': [[item.get('HK')]]},
                {'range': f'J{next_row}', 'values': [[item.get('No Pengajuan')]]},
                {'range': f'K{next_row}', 'values': [[item.get('Harga Cust')]]},
                {'range': f'L{next_row}', 'values': [[item.get('gap')]]},
                {'range': f'M{next_row}', 'values': [[item.get('persen')]]},
                {'range': f'N{next_row}', 'values': [[item.get('potensi')]]},
                {'range': f'O{next_row}', 'values': [[item.get('support')]]},
                {'range': f'P{next_row}', 'values': [[item.get('rasio')]]},
                {'range': f'Q{next_row}', 'values': [[item.get('Keterangan')]]}
            ]
            sheet.batch_update(updates)
            next_row += 1 
    except Exception as e:
        st.error(f"Gagal simpan ke Sheets: {e}")
        
# --- UI INPUT ---
st.title("Form Pengajuan Voucher Big Sales")
with st.container(border=True):
    col1, col2 = st.columns(2)
    region = col1.selectbox("Region", list(data_toko.keys()))
    store = col2.selectbox("Store", data_toko[region])
with st.container(border=True):
    col3, col4 = st.columns(2)
    no_cust = col3.text_input("No. Customer")
    nama_cust = col4.text_input("Nama Customer")

# --- KOTAK 3: INPUT PRODUK ---
with st.container(border=True):
    st.write("##### 3. Input Produk")
    input_barcode = st.text_input("Masukkan Barcode")
    auto_prodname = ""
    
    if input_barcode:
        match = product_db[product_db['Barcode'] == input_barcode]
        if not match.empty:
            auto_prodname = match.iloc[0]['Prodname']
            st.success(f"Produk ditemukan: {auto_prodname}")
        else:
            st.warning("Barcode tidak ditemukan.")

    with st.form("input_form", clear_on_submit=True):
        c1, c2 = st.columns(2)
        barcode = c1.text_input("Barcode", value=input_barcode, disabled=True)
        prodname = c2.text_input("Prodname", value=auto_prodname, disabled=True)
        
        c3, c4, c5 = st.columns(3)
        qty = c3.number_input("QTY", min_value=0, step=1, format="%d")
        hk_buyer = c4.number_input("HK Buyer", min_value=0, step=1, format="%d")
        harga_cust = c5.number_input("Harga Customer", min_value=0, step=1, format="%d")
        no_pengajuan = st.text_input("No Pengajuan HK Buyer")
        ket = st.text_input("Keterangan")
        submitted = st.form_submit_button("Cek Perhitungan")

    if submitted:
        st.session_state.temp = {
            "barcode": input_barcode,
            "prodname": auto_prodname,
            "gap": hk_buyer - harga_cust,
            "persen": ((hk_buyer - harga_cust) / hk_buyer * 100) if hk_buyer != 0 else 0,
            "potensi": qty * hk_buyer,
            "support": qty * (hk_buyer - harga_cust),
            "rasio": ((qty * (hk_buyer - harga_cust)) / (qty * hk_buyer) * 100) if (qty * hk_buyer) != 0 else 0
        }

    if 'temp' in st.session_state:
        m1, m2, m3 = st.columns(3)
        m1.metric("Gap (Value)", f"{st.session_state.temp['gap']:,.0f}")
        m2.metric("Gap (%)", f"{st.session_state.temp['persen']:,.2f}%")
        m3.metric("Potensi Sales", f"{st.session_state.temp['potensi']:,.0f}")
        m4, m5 = st.columns(2)
        m4.metric("Support Voucher", f"{st.session_state.temp['support']:,.0f}")
        m5.metric("Rasio Voucher", f"{st.session_state.temp['rasio']:,.2f}%")
        
        if st.button("Tambah ke Daftar"):
            # Gabungkan langsung, tidak perlu membuat dua kali
            item_data = {
                "No Cust": no_cust, 
                "Nama": nama_cust, 
                "Barcode": input_barcode, # Gunakan input_barcode langsung
                "Prodname": auto_prodname, # Gunakan auto_prodname langsung
                "QTY": qty, 
                "HK": hk_buyer, 
                "Harga Cust": harga_cust, 
                "No Pengajuan": no_pengajuan, 
                "Keterangan": ket
            }
            # Tambahkan data perhitungan dari temp
            item_data.update(st.session_state.temp)
            
            # Hapus key yang duplikat jika ada (opsional, agar bersih)
            if 'barcode' in item_data: del item_data['barcode']
            if 'prodname' in item_data: del item_data['prodname']
            
            st.session_state.daftar_pengajuan.append(item_data)
            
            del st.session_state.temp
            st.rerun()

# --- DAFTAR & GENERATE ---
if st.session_state.daftar_pengajuan:
    st.write("### Daftar Pengajuan")
    st.table(pd.DataFrame(st.session_state.daftar_pengajuan))
    
    if st.button("Generate & Download Excel"):
        try:
            simpan_ke_googlesheets(st.session_state.daftar_pengajuan, store)
            
            wb = openpyxl.load_workbook("VCR TEMPLATE.xlsx")
            ws = wb.active
            center_align = Alignment(horizontal='center', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')

            nama_ttd = {"Region 1": "AZKAHADI PUTRA", "Region 2": "DESKY BAYU AJI", "Region 3": "ADE CHANDRA"}
            ws.cell(row=9, column=13, value=nama_ttd.get(region, ""))
            ws.cell(row=9, column=13).alignment = center_align
            
            for merged_range in list(ws.merged_cells.ranges):
                if merged_range.min_row >= 14: ws.unmerge_cells(str(merged_range))

            formats = {2: '@', 3: '@', 4: '@', 5: '@', 6: '0', 7: '#,##0', 8: '@', 9: '#,##0', 10: '#,##0', 12: '#,##0', 13: '#,##0'}
            for i, item in enumerate(st.session_state.daftar_pengajuan):
                row = 14 + i
                data_row = [i + 1, str(item.get('No Cust', '')), str(item.get('Nama', '')), str(item.get('Barcode', '')), 
                            str(item.get('Prodname', '')), int(item.get('QTY', 0)), int(item.get('HK', 0)), 
                            str(item.get('No Pengajuan', '')), int(item.get('Harga Cust', 0)), int(item.get('gap', 0)), 
                            item.get('persen', 0) / 100, int(item.get('potensi', 0)), int(item.get('support', 0)), 
                            item.get('rasio', 0) / 100, str(item.get('Keterangan', ''))]
                
                for col_idx, value in enumerate(data_row, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.alignment = left_align if col_idx in [3, 5] else center_align
                    if col_idx in [11, 14]: cell.number_format = '0.00%'
                    elif col_idx in formats: cell.number_format = formats[col_idx]

            buffer = BytesIO()
            wb.save(buffer)
            st.download_button("Download Excel Rapi", buffer.getvalue(), f"Pengajuan_{store}.xlsx")
            st.success("Data tersimpan ke Sheets & Excel siap!")
        except Exception as e:
            st.error(f"Error: {e}")

    if st.button("Reset Daftar"):
        st.session_state.daftar_pengajuan = []
        st.rerun()
